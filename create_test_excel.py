"""
Create comprehensive test Excel file for multi-column time series feature
Clean format: Column names in first row, no extra instructions in sheets
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Create a new workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center_align = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_sheet(ws, df):
    """Apply styling to worksheet - clean format with headers in first row"""
    # Add data with headers directly (no title, no extra rows)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = center_align
            
            # Header row styling (first row only)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================================
# TEST CASE 1: Numeric Week Numbers (Single Column)
# ============================================================================
print("Creating Test Case 1: Numeric Week Numbers...")
ws1 = wb.create_sheet("WeekNumbers")

data1 = {
    'Weeknumber': [12301, 12302, 12303, 12304, 12305, 12306, 12307, 12308, 12309, 12310,
                   12311, 12312, 12313, 12314, 12315, 12316, 12317, 12318, 12319, 12320],
    'Sales': [1250, 1380, 1420, 1150, 1680, 1520, 1390, 1450, 1280, 1610,
              1550, 1720, 1340, 1480, 1590, 1410, 1530, 1670, 1290, 1440],
    'Region': ['North'] * 10 + ['South'] * 10,
    'Defects': [2, 1, 3, 5, 0, 2, 4, 1, 3, 2, 1, 0, 2, 3, 1, 4, 2, 1, 5, 3]
}
df1 = pd.DataFrame(data1)
style_sheet(ws1, df1)

# ============================================================================
# TEST CASE 2: Year + Month (Numeric Combination)
# ============================================================================
print("Creating Test Case 2: Year + Month...")
ws2 = wb.create_sheet("Year_Month")

np.random.seed(42)
data2 = {
    'Year': [2023] * 12 + [2024] * 8,
    'Month': list(range(1, 13)) + list(range(1, 9)),
    'Sales': np.random.randint(1000, 2000, 20),
    'Region': ['North', 'South'] * 10,
    'Defects': np.random.randint(0, 6, 20)
}
df2 = pd.DataFrame(data2)
style_sheet(ws2, df2)

# ============================================================================
# TEST CASE 3: Date + Time
# ============================================================================
print("Creating Test Case 3: Date + Time...")
ws3 = wb.create_sheet("Date_Time")

dates = pd.date_range('2023-01-01', periods=15, freq='D')
times = ['09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', 
         '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00']

data3 = {
    'Date': [d.strftime('%Y-%m-%d') for d in dates],
    'Time': times,
    'Temperature': [22.5, 23.1, 24.3, 25.7, 26.2, 27.1, 28.3, 27.8, 26.9, 25.4, 24.6, 23.8, 23.2, 22.9, 22.4],
    'Sensor': ['A'] * 8 + ['B'] * 7,
    'Alerts': [0, 0, 0, 1, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0]
}
df3 = pd.DataFrame(data3)
style_sheet(ws3, df3)

# ============================================================================
# TEST CASE 4: Year + Quarter
# ============================================================================
print("Creating Test Case 4: Year + Quarter...")
ws4 = wb.create_sheet("Year_Quarter")

data4 = {
    'Year': [2022] * 4 + [2023] * 4 + [2024] * 4,
    'Quarter': [1, 2, 3, 4] * 3,
    'Revenue': [45000, 52000, 48000, 61000, 47000, 55000, 51000, 64000, 49000, 58000, 53000, 67000],
    'Product': ['Widget'] * 6 + ['Gadget'] * 6,
    'Returns': [120, 98, 145, 87, 132, 105, 156, 92, 128, 110, 148, 95]
}
df4 = pd.DataFrame(data4)
style_sheet(ws4, df4)

# ============================================================================
# TEST CASE 5: Incompatible Columns
# ============================================================================
print("Creating Test Case 5: Incompatible columns...")
ws5 = wb.create_sheet("Incompatible")

data5 = {
    'Region': ['North', 'South', 'East', 'West', 'North', 'South', 'East', 'West',
               'North', 'South', 'East', 'West', 'North', 'South', 'East'],
    'Product': ['WidgetA', 'WidgetB', 'GadgetX', 'GadgetY', 'WidgetA', 'WidgetB', 'GadgetX', 'GadgetY',
                'WidgetA', 'WidgetB', 'GadgetX', 'GadgetY', 'WidgetA', 'WidgetB', 'GadgetX'],
    'Sales': np.random.randint(800, 1500, 15),
    'Defects': np.random.randint(0, 5, 15)
}
df5 = pd.DataFrame(data5)
style_sheet(ws5, df5)

# ============================================================================
# TEST CASE 6: Unsorted Data
# ============================================================================
print("Creating Test Case 6: Unsorted data...")
ws6 = wb.create_sheet("Unsorted_Weeks")

data6 = {
    'Weeknumber': [12305, 12301, 12308, 12302, 12310, 12304, 12307, 12303, 12309, 12306,
                   12315, 12312, 12318, 12314, 12311, 12320, 12316, 12313, 12319, 12317],
    'Sales': [1680, 1250, 1450, 1380, 1610, 1150, 1390, 1420, 1280, 1520,
              1590, 1720, 1670, 1480, 1550, 1440, 1530, 1340, 1290, 1410],
    'Region': ['North'] * 10 + ['South'] * 10,
    'Defects': [0, 2, 1, 1, 2, 5, 4, 3, 3, 2, 1, 0, 1, 3, 1, 3, 2, 2, 5, 4]
}
df6 = pd.DataFrame(data6)
style_sheet(ws6, df6)

# ============================================================================
# TEST CASE 7: Year + Month + Day
# ============================================================================
print("Creating Test Case 7: Year + Month + Day...")
ws7 = wb.create_sheet("Year_Month_Day")

data7 = {
    'Year': [2023] * 15,
    'Month': [1, 1, 1, 2, 2, 2, 3, 3, 3, 4, 4, 4, 5, 5, 5],
    'Day': [5, 12, 19, 2, 16, 23, 9, 16, 23, 6, 13, 20, 4, 11, 18],
    'Value': [145, 158, 167, 142, 171, 156, 163, 149, 174, 138, 168, 155, 162, 151, 169],
    'Category': ['A'] * 5 + ['B'] * 5 + ['A'] * 5
}
df7 = pd.DataFrame(data7)
style_sheet(ws7, df7)

# ============================================================================
# TEST CASE 8: Standard DateTime Column
# ============================================================================
print("Creating Test Case 8: Standard datetime...")
ws8 = wb.create_sheet("Standard_Date")

dates = pd.date_range('2023-01-01', periods=20, freq='W')
data8 = {
    'Date': [d.strftime('%Y-%m-%d') for d in dates],
    'Sales': np.random.randint(1000, 2000, 20),
    'Store': ['Store_A'] * 10 + ['Store_B'] * 10,
    'Defects': np.random.randint(0, 5, 20)
}
df8 = pd.DataFrame(data8)
style_sheet(ws8, df8)

# ============================================================================
# Add README Sheet at the beginning
# ============================================================================
print("Creating README sheet...")
ws_readme = wb.create_sheet("README", 0)

title_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=14)
section_font = Font(bold=True, size=12, color="2E75B6")

readme_content = [
    ["DART ANALYTICS - MULTI-COLUMN TIME SERIES TEST FILE"],
    [""],
    ["OVERVIEW"],
    ["This file contains 8 test cases for the multi-column time series feature."],
    ["Each sheet is ready to use with column names in the first row."],
    [""],
    ["TEST CASES"],
    [""],
    ["Sheet", "Description", "Time Series Columns", "Value Column"],
    ["WeekNumbers", "Single numeric week column", "Weeknumber", "Sales"],
    ["Year_Month", "Combine Year + Month", "Year + Month (check both)", "Sales"],
    ["Date_Time", "Combine Date + Time", "Date + Time (check both)", "Temperature"],
    ["Year_Quarter", "Combine Year + Quarter", "Year + Quarter (check both)", "Revenue"],
    ["Incompatible", "Error test - incompatible columns", "Region + Product (should error)", "Sales"],
    ["Unsorted_Weeks", "Auto-sort test", "Weeknumber (unsorted)", "Sales"],
    ["Year_Month_Day", "Three column combination", "Year + Month + Day (all 3)", "Value"],
    ["Standard_Date", "Normal date column", "Date", "Sales"],
    [""],
    ["HOW TO USE"],
    [""],
    ["1. Upload this file to DART Analytics"],
    ["2. Select the sheet you want to test"],
    ["3. Select Value column from dropdown"],
    ["4. Check one or more Time Series column checkboxes"],
    ["5. Click Generate Charts"],
    [""],
    ["MULTI-COLUMN FEATURE"],
    [""],
    ["• Check multiple checkboxes to combine columns"],
    ["• Example: Check Year + Month to create combined time series"],
    ["• System automatically combines compatible columns"],
    ["• Numeric columns: concatenated (2023 + 1 = 20231)"],
    ["• Date/Time columns: parsed to datetime"],
    [""],
    ["EXPECTED RESULTS"],
    [""],
    ["✓ WeekNumbers: Single column analysis"],
    ["✓ Year_Month: Combines to 20231, 20232, etc."],
    ["✓ Date_Time: Combines to full datetime"],
    ["✓ Year_Quarter: Combines to 20221, 20222, etc."],
    ["✗ Incompatible: Should show error message"],
    ["✓ Unsorted_Weeks: Auto-sorts before analysis"],
    ["✓ Year_Month_Day: Combines all three columns"],
    ["✓ Standard_Date: Normal single date analysis"],
]

for row_idx, row_data in enumerate(readme_content, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_readme.cell(row=row_idx, column=col_idx, value=value)
        
        # Title row
        if row_idx == 1:
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = center_align
        # Section headers
        elif value in ["OVERVIEW", "TEST CASES", "HOW TO USE", "MULTI-COLUMN FEATURE", "EXPECTED RESULTS"]:
            cell.font = section_font
        # Table header
        elif row_idx == 9:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Merge title
ws_readme.merge_cells('A1:D1')

# Adjust column widths for README
ws_readme.column_dimensions['A'].width = 20
ws_readme.column_dimensions['B'].width = 40
ws_readme.column_dimensions['C'].width = 35
ws_readme.column_dimensions['D'].width = 20

# Save the workbook
output_file = "DART_MultiColumn_TimeSeries_TestCases.xlsx"
wb.save(output_file)
print(f"\n✅ Test Excel file created successfully: {output_file}")
print(f"📊 Contains 8 test sheets + README")
print(f"🎯 Clean format: Column names in first row")
print(f"🚀 Ready to upload to DART Analytics!\n")
